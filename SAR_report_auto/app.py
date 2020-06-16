import sys,UI
import openpyxl as op
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIcon,QFont
from PyQt5.QtCore import QCoreApplication,QDate,Qt
from PyQt5 import uic


class MyApp(UI.Ui_Dialog,QDialog):


    def __init__(self):
        super().__init__()
        QDialog.__init__(self, None, Qt.WindowStaysOnTopHint)
        self.setupUi(self)
        self.date = QDate.currentDate()
        self.initUI()

    def initUI(self):

        QToolTip.setFont(QFont('SansSerif', 10))
        self.center()

        #버튼과 ui 상호작용
        self.browse_btn.clicked.connect(self.OnOpenDocument1)
        self.browse_btn.setToolTip('파일 불러오기')


        self.quit_btn.clicked.connect(QCoreApplication.instance().quit)
        self.quit_btn.setToolTip('종료하기2')

        self.OK_btn.clicked.connect(self.okFunction)
        self.OK_btn.setToolTip('확인')

        self.setWindowTitle('SAR Report Assistant')
        
        self.resize(500, 225)
        self.show()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())


    def OnOpenDocument1(self):
        fname1 = QFileDialog.getOpenFileName(self, 'Open file', "",
                                            "All Files(*);; Python Files(*.py);; CSV Files(*.csv);; Excel Files(*.xlsx)", '/home')
        global send_name
        send_name = fname1[0]
        self.fileBox.setText(fname1[0])
        print(send_name)

    



    def okFunction(self):
        def merge_cells(st,end,line):
                input_string = ''
                input_string += str(line)
                input_string += str(st)
                input_string += ':'
                input_string += str(line)
                input_string += str(end)
                sheet.merge_cells(input_string)


        file_name = op.load_workbook(send_name)
        sheet = file_name['Sheet1']

        line = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
        line_num = 0

        for x in line:
            if (sheet[x][19].value == None) and (sheet[x][20].value == None):
                break
            line_num+=1

        switch_num = 0

        for x  in range(line_num):
            for y in range(19,len(sheet['A'])):

                if sheet[line[x]][y].value == None and y==19:
                    sheet[line[x]][y].value = sheet[line[x-1]][y].value
                elif sheet[line[x]][y].value == None and y==20:
                    sheet[line[x]][y].value = sheet[line[x]][y-1].value
                elif y>20 and sheet[line[x]][y].value==None and switch_num==0:
                    sheet[line[x]][y].value = sheet[line[x]][y-1].value
                elif switch_num==1 and y>20:
                    sheet[line[x]][y].value = sheet[line[x]][y].value
            
            if sheet[line[x]][19].value == 'Freq. (MHz)':
                switch_num=1

        n = 19

        while n<len(sheet['A']):
            if sheet[line[line_num-2]][n].value == None:
                sheet.delete_rows(n+1)
                n=19
            
            n+=1
            if n==1000:
                break

        for x in range(line_num):
            st=19
            last = sheet[line[x]][19].value
            if line[x]=='A':
                for y in range(19,len(sheet['A'])):
                    #print('y is : ',y , '[a] : ', len(sheet['A']))
                    
                    if last != sheet[line[x]][y].value or y==(len(sheet['A'])-1):
                        #print(st+1,y,line[x])
                        merge_cells(st+1,y,line[x])
                        merge_cells(st+1,y,line[x+1])
                        merge_cells(st+1,y,line[x+2])
                        merge_cells(st+1,y,line[x+3])
                        last = sheet[line[x]][y].value
                        st = y
            elif line[x] == 'E':
                for y in range(19,len(sheet['A'])):
                    if last != sheet[line[x]][y].value or y==(len(sheet['A'])-1):
                        #print(st+1,y,line[x])
                        merge_cells(st+1,y,line[x])
                        last = sheet[line[x]][y].value
                        st = y


        file_name.save('result_file.xlsx')

if __name__ == '__main__':
   app = QApplication(sys.argv)
   ex = MyApp()
   sys.exit(app.exec_())
