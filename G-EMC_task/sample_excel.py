import openpyxl as op
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, Protection
from openpyxl import Workbook, load_workbook
import time

def sample_excel(user_defined_Fname,model_name,model_id,user_name,created):
    #workbook 생성
    wb = Workbook()
    #워크북을 생성하면 적어도 하나의 워크 시트를 생성해야 한다.
    sht = wb.active
    sht.title = model_name # 현재 접속중인 모델명을 가져다 넣는다.

    center = Alignment(horizontal='center', vertical = 'center')
    #thin_border = Border(left=Side(style='middle'), right=Side(style='middle'), top=Side(style='middle'), bottom=Side(style='middle'))
    # 선택되는 process에는 색칠하기.

    #테두리 그리기
    sht['B2'].border = Border(left=Side(style='medium'), top=Side(style='medium'))
    sht['B3'].border = Border(left=Side(style='medium'))
    sht['B4'].border = Border(left=Side(style='medium'))
    sht['B5'].border = Border(left=Side(style='medium'),bottom=Side(style='medium'))

    sht['C2'].border = Border(top=Side(style='medium'))
    sht['C5'].border = Border(bottom=Side(style='medium'))

    sht['D2'].border = Border(top=Side(style='medium'), right=Side(style='medium'))
    sht['D3'].border = Border(right=Side(style='medium'))
    sht['D4'].border = Border(right=Side(style='medium'))
    sht['D5'].border = Border(right=Side(style='medium'), bottom=Side(style='medium'))

    sht['B7'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'))
    sht['B8'].border = Border(left=Side(style='medium'),right=Side(style='medium'),bottom=Side(style='medium'))
    sht['B9'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C9'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C12'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C15'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C18'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))

    sht['D9'].border = Border(right=Side(style='medium'),top=Side(style='medium'))
    sht['D10'].border = Border(right=Side(style='medium'))
    sht['D11'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D12'].border = Border(right=Side(style='medium'))
    sht['D13'].border = Border(right=Side(style='medium'))
    sht['D14'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D15'].border = Border(right=Side(style='medium'))
    sht['D16'].border = Border(right=Side(style='medium'))
    sht['D17'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D18'].border = Border(right=Side(style='medium'))
    sht['D19'].border = Border(right=Side(style='medium'))
    sht['D20'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))

    sht['B22'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C22'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C25'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C28'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C31'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C34'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C37'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))
    sht['C40'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'),bottom=Side(style='medium'))

    sht['D22'].border = Border(right=Side(style='medium'),top=Side(style='medium'))
    sht['D23'].border = Border(right=Side(style='medium'))
    sht['D24'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D25'].border = Border(right=Side(style='medium'))
    sht['D26'].border = Border(right=Side(style='medium'))
    sht['D27'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D28'].border = Border(right=Side(style='medium'))
    sht['D29'].border = Border(right=Side(style='medium'))
    sht['D30'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D31'].border = Border(right=Side(style='medium'))
    sht['D32'].border = Border(right=Side(style='medium'))
    sht['D33'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D34'].border = Border(right=Side(style='medium'))
    sht['D35'].border = Border(right=Side(style='medium'))
    sht['D36'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D37'].border = Border(right=Side(style='medium'))
    sht['D38'].border = Border(right=Side(style='medium'))
    sht['D39'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    sht['D40'].border = Border(right=Side(style='medium'))
    sht['D41'].border = Border(right=Side(style='medium'))
    sht['D42'].border = Border(right=Side(style='medium'),bottom=Side(style='medium'))

    #기본정보 입력칸
    sht.merge_cells('B2:C2')
    sht['B2'] = '제품명'
    sht['B2'].font = Font(bold=True)
    sht['D2'] = model_name # 입력으로 받기
    sht['B2'].alignment = center
    sht['D2'].alignment = center


    sht.merge_cells('B3:C3')
    sht['B3'] = '모델명'
    sht['B3'].font = Font(bold=True)
    sht['D3'] = model_id # 모델명은 자동으로 입력 (형광펜 밑줄 VALUE)
    sht['B3'].alignment = center
    sht['D3'].alignment = center

    sht.merge_cells('B4:C4')
    sht['B4'] = '담당자'
    sht['B4'].font = Font(bold=True)
    sht['D4'] = user_name # 현재 USER_NAME 가져오기
    sht['B4'].alignment = center
    sht['D4'].alignment = center

    sht.merge_cells('B5:C5')
    sht['B5'] = '프로젝트 시작일'
    sht['B5'].font = Font(bold=True)
    sht['D5'] = created # tbl_wish 만든 날짜 가져오기
    sht['B5'].alignment = center
    sht['D5'].alignment = center

    # 예상 프로젝트 완료일을 엑셀에 작성할지는 일단 보류

    sht.merge_cells('B7:D7')
    sht.merge_cells('B8:D8')
    #sht.merge_cells('B7:B8')
    sht['B8'] = '시험항목'
    sht['B8'].font = Font(bold=True)
    sht['B8'].alignment = center

    sht.merge_cells('B9:B20')
    sht['B9'] = 'EMI'
    sht['B9'].alignment = center

    sht.merge_cells('C9:C11')
    sht['C9'] = 'RE'
    sht['D9'] = '120V, 60 Hz'
    sht['D10'] = '220V, 60 Hz'
    sht['D11'] = '230V, 50 Hz'
    sht['C9'].alignment = center
    sht['D9'].alignment = center
    sht['D10'].alignment = center
    sht['D11'].alignment = center

    sht.merge_cells('C12:C14')
    sht['C12'] = 'CE'
    sht['D12'] = '120V, 60 Hz'
    sht['D13'] = '220V, 60 Hz'
    sht['D14'] = '230V, 50 Hz'
    sht['C12'].alignment = center
    sht['D12'].alignment = center
    sht['D13'].alignment = center
    sht['D14'].alignment = center

    sht.merge_cells('C15:C17')
    sht['C15'] = 'Harmonic'
    sht['D15'] = '120V, 60 Hz'
    sht['D16'] = '220V, 60 Hz'
    sht['D17'] = '230V, 50 Hz'
    sht['C15'].alignment = center
    sht['D15'].alignment = center
    sht['D16'].alignment = center
    sht['D17'].alignment = center

    sht.merge_cells('C18:C20')
    sht['C18'] = 'Flicker'
    sht['D18'] = '120V, 60 Hz'
    sht['D19'] = '220V, 60 Hz'
    sht['D20'] = '230V, 50 Hz'
    sht['C18'].alignment = center
    sht['D18'].alignment = center
    sht['D19'].alignment = center
    sht['D20'].alignment = center

    sht.merge_cells('B22:B42')
    sht['B22'] = 'EMS'
    sht['B22'].alignment = center

    sht.merge_cells('C22:C24')
    sht['C22'] = 'ESD'
    sht['D22'] = '120V, 60 Hz'
    sht['D23'] = '220V, 60 Hz'
    sht['D24'] = '230V, 50 Hz'
    sht['C22'].alignment = center
    sht['D22'].alignment = center
    sht['D23'].alignment = center
    sht['D24'].alignment = center

    sht.merge_cells('C25:C27')
    sht['C25'] = 'RS'
    sht['D25'] = '120V, 60 Hz'
    sht['D26'] = '220V, 60 Hz'
    sht['D27'] = '230V, 50 Hz'
    sht['C25'].alignment = center
    sht['D25'].alignment = center
    sht['D26'].alignment = center
    sht['D27'].alignment = center

    sht.merge_cells('C28:C30')
    sht['C28'] = 'EFT/Burst'
    sht['D28'] = '120V, 60 Hz'
    sht['D29'] = '220V, 60 Hz'
    sht['D30'] = '230V, 50 Hz'
    sht['C28'].alignment = center
    sht['D28'].alignment = center
    sht['D29'].alignment = center
    sht['D30'].alignment = center

    sht.merge_cells('C31:C33')
    sht['C31'] = 'Surge'
    sht['D31'] = '120V, 60 Hz'
    sht['D32'] = '220V, 60 Hz'
    sht['D33'] = '230V, 50 Hz'
    sht['C31'].alignment = center
    sht['D31'].alignment = center
    sht['D32'].alignment = center
    sht['D33'].alignment = center

    sht.merge_cells('C34:C36')
    sht['C34'] = 'CS'
    sht['D34'] = '120V, 60 Hz'
    sht['D35'] = '220V, 60 Hz'
    sht['D36'] = '230V, 50 Hz'
    sht['C34'].alignment = center
    sht['D34'].alignment = center
    sht['D35'].alignment = center
    sht['D36'].alignment = center

    sht.merge_cells('C37:C39')
    sht['C37'] = 'MF'
    sht['D37'] = '120V, 60 Hz'
    sht['D38'] = '220V, 60 Hz'
    sht['D39'] = '230V, 50 Hz'
    sht['C37'].alignment = center
    sht['D37'].alignment = center
    sht['D38'].alignment = center
    sht['D39'].alignment = center

    sht.merge_cells('C40:C42')
    sht['C40'] = 'V-Dip/Interuption'
    sht['D40'] = '120V, 60 Hz'
    sht['D41'] = '220V, 60 Hz'
    sht['D42'] = '230V, 50 Hz'
    sht['C40'].alignment = center
    sht['D40'].alignment = center
    sht['D41'].alignment = center
    sht['D42'].alignment = center

    # 행 너비 설정
    sht.row_dimensions[8].height = 50

    sht.column_dimensions['B'].width = 10
    sht.column_dimensions['C'].width = 20
    sht.column_dimensions['D'].width = 30

    #보기 좋게
    sht.merge_cells('B6:C6')
    #sht.merge_cells('B7:C7')

    # 동그라미랑 이제 날짜를 입력해 보자.
    # 그 전에 모드도 입력해야함. B7: Regulation B8: Mode _name
    File_name = user_defined_Fname + '.xlsx'
    wb.save(File_name) # SM-8498 이런식으로 저장
    print('finiszzzzzh')


def add_excel (user_defined_Fname,count,r,v):
    File_name = user_defined_Fname + '.xlsx'

    load_wb = load_workbook(File_name)
    print("> add_excel ()", File_name)
    load_ws = load_wb[user_defined_Fname]

    #print("---------------------------------------------------------------------",last_col)
    last_col = ord('E')+2*(count-1)

    r_comb = chr(last_col) + str(7)
    v_comb = chr(last_col) + str(8)
    load_ws[r_comb].value =r
    load_ws[v_comb].value =v

    center = Alignment(horizontal='center', vertical = 'center')
    load_ws[r_comb].alignment = center
    load_ws[v_comb].alignment = center

    load_ws.column_dimensions[chr(last_col)].width = 30
    load_ws.row_dimensions[7].height = 50

    load_wb.save(File_name)
    return


def update_excel(user_defined_Fname,count,r,v,check_array):
    dic_array = {'RE':9, 'CE':12, 'Harmonic':15,'Flicker':18,'ESD':22, 'RS':25, 'EFT/Burst':28, 'Surge':31, 'CS':34, 'MF':37, 'V-Dip/Interuption':40}
    search_Scol = 'E' # + count 만큼 탐색!
    #E7 : Regulation, E8: mode

    File_name = user_defined_Fname + '.xlsx'

    load_wb = load_workbook(File_name)
    load_ws = load_wb[user_defined_Fname]
    print("------ update_excel()-----------------------")
    center = Alignment(horizontal='center', vertical = 'center')

    for c in range(0,count):
        r_comb = chr(ord(search_Scol)+2*c) + str(7)
        v_comb = chr(ord(search_Scol)+2*c) + str(8)
        print(r_comb)

        print("yyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyy",load_ws[r_comb].value)

        # 해당 규격/ 모드명에 도달했을 때.
        if load_ws[r_comb].value==r and load_ws[v_comb].value==v:
            # 동그라미 고고
            for i in check_array:
                print("forfor",dic_array[i])
                start_row = dic_array[i]
                #start_col = 5 #'E'
                for j in range(3):
                    row = start_row + j
                    p = chr(ord(search_Scol) + 2*(c)) + str(row)
                    p_next = chr(ord(search_Scol) + 2*(c) +1) + str(row)
                    load_ws[p].value = 'O'
                    dt = time.strftime('%m/%d', time.localtime(time.time()))
                    load_ws[p_next].value = dt

                    load_ws[p].alignment = center
                    load_ws[p_next].alignment = center
            break

    print("update_excel 성공")

    load_wb.save(File_name)
    return
