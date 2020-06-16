import sys,UI
import openpyxl as op
from openpyxl.styles import  PatternFill,Color,Font, Alignment, Border, Side
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIcon,QFont
from PyQt5.QtCore import QCoreApplication,QDate,Qt
from collections import Counter
import copy
import re
import pandas as pd
from PyQt5 import uic
sys.setrecursionlimit(5000)

class MyApp(UI.Ui_Dialog,QDialog):


    def __init__(self):
        super().__init__()
        QDialog.__init__(self, None, Qt.WindowStaysOnTopHint)
        self.setupUi(self)
        self.date = QDate.currentDate()
        self.initUI()

    def initUI(self):
        QToolTip.setFont(QFont('SansSerif', 10))
        self.center()

        #버튼과 ui 상호작용
        self.browse1.clicked.connect(self.OnOpenDocument1)#fname1
        self.browse1.setToolTip('파일 불러오기1 입니다.')

        self.browse2.clicked.connect(self.OnOpenDocument2)
        self.browse2.setToolTip('파일 불러오기2 입니다.')

        self.browse3.clicked.connect(self.OnOpenDocument3)
        self.browse3.setToolTip('파일 불러오기3 입니다.')

        self.browse4.clicked.connect(self.OnOpenDocument4)
        self.browse4.setToolTip('파일 불러오기4 입니다.')

#-----------------SISO + MIMO table 추가 구현 ------------------------------

        self.browse0.clicked.connect(self.OnOpenDocument0)
        self.browse0.setToolTip('파일 불러오기0 입니다.')

        self.browse_1.clicked.connect(self.OnOpenDocument_1)
        self.browse_1.setToolTip('파일 불러오기_1 입니다.')

#----------------------------------------------------------------
        self.cancel2.clicked.connect(QCoreApplication.instance().quit)
        self.cancel2.setToolTip('종료하기2')

        self.cancel1.clicked.connect(QCoreApplication.instance().quit)
        self.cancel1.setToolTip('종료하기1')

        self.cancel1.clicked.connect(QCoreApplication.instance().quit)
        self.cancel1.setToolTip('종료하기0')


        self.ok1.clicked.connect(self.okFunction)
        self.ok1.setToolTip('확인1')

        self.ok2.clicked.connect(self.okFunction2)
        self.ok2.setToolTip('확인2')

#--------------SISO + MIMO table 추가 구현 ---------------------------------

        self.ok0.clicked.connect(self.okFunction0)
        self.ok0.setToolTip('확인0')



#######################################################################
#상단 위도우 박스 정리 및 크기 레이아웃 정리
#로고 정리

        self.setWindowTitle('SAR Team CA Table Assistant')
        self.move(600, 600)
        self.resize(900, 600)
        self.show()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())


    def OnOpenDocument1(self):
        fname1 = QFileDialog.getOpenFileName(self, 'Open file', "",
                                            "All Files(*);; Python Files(*.py);; CSV Files(*.csv);; Excel Files(*.xlsx)", '/home')

        global send_name1
        send_name1 = fname1[0]
        self.fname1.setText(fname1[0])
        print(send_name1)

    def OnOpenDocument2(self):
        fname2 = QFileDialog.getOpenFileName(self, 'Open file', "",
                                            "All Files(*);; Python Files(*.py);; CSV Files(*.csv);; Excel Files(*.xlsx)", '/home')

        global send_name2
        send_name2 = fname2[0]
        self.fname2.setText(fname2[0])
        print(send_name2)

    def OnOpenDocument3(self):
        fname3 = QFileDialog.getOpenFileName(self, 'Open file', "",
                                            "All Files(*);; Python Files(*.py);; CSV Files(*.csv);; Excel Files(*.xlsx)", '/home')

        global send_name3
        send_name3 = fname3[0]
        self.fname3.setText(fname3[0])
        print(send_name3)

    def OnOpenDocument4(self):
        fname4 = QFileDialog.getOpenFileName(self, 'Open file', "",
                                            "All Files(*);; Python Files(*.py);; CSV Files(*.csv);; Excel Files(*.xlsx)", '/home')

        global send_name4
        send_name4 = fname4[0]
        self.fname4.setText(fname4[0])
        print(send_name4)

    def OnOpenDocument0(self):
        fname0 = QFileDialog.getOpenFileName(self, 'Open file', "",
                                            "All Files(*);; Python Files(*.py);; CSV Files(*.csv);; Excel Files(*.xlsx)", '/home')

        global send_name0
        send_name0 = fname0[0]
        self.fname0.setText(fname0[0])
        print(send_name0)

    def OnOpenDocument_1(self):
        fname_1 = QFileDialog.getOpenFileName(self, 'Open file', "",
                                            "All Files(*);; Python Files(*.py);; CSV Files(*.csv);; Excel Files(*.xlsx)", '/home')

        global send_name_1
        send_name_1 = fname_1[0]
        self.fname_1.setText(fname_1[0])
        print(send_name_1)


    def okFunction0(self):
        ###엑셀에서 불러오는 시트는 무조건 combo 일것
        divide =  op.load_workbook(send_name0)
        divide_sheet = divide['combo']
        basic = op.load_workbook(send_name_1)
        basic_sheet = basic['Sheet1']
        basic_sheet2 = basic['Sheet2']

        len_cc = len(divide_sheet['B'])

        cc1_num = 1
        cc2_num = 1
        cc3_num = 1
        cc4_num = 1
        cc5_num = 1
        cc1_num2 = 1
        cc2_num2 = 1
        cc3_num2 = 1
        cc4_num2 = 1
        cc5_num2 = 1


        for x in range(3,len_cc):
            if divide_sheet['B'][x].value == None:
                break

            if '[' not in list(divide_sheet['C'][x].value):
                if divide_sheet['B'][x].value[0] == '1':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #' + str(cc1_num)
                    basic_sheet['A'][int(cc1_num)+2].value = str_cc
                    basic_sheet['B'][int(cc1_num)+2].value = divide_sheet['C'][x].value
                    basic_sheet['E'][int(cc1_num)+2].value = 'YES'
                    cc1_num+=1
                elif divide_sheet['B'][x].value[0] == '2':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #' + str(cc2_num)
                    basic_sheet['F'][int(cc2_num)+2].value = str_cc
                    basic_sheet['G'][int(cc2_num)+2].value = divide_sheet['C'][x].value
                    basic_sheet['J'][int(cc2_num)+2].value = 'YES'
                    cc2_num+=1
                elif divide_sheet['B'][x].value[0] == '3':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #' + str(cc3_num)
                    basic_sheet['K'][int(cc3_num)+2].value = str_cc
                    basic_sheet['L'][int(cc3_num)+2].value = divide_sheet['C'][x].value
                    basic_sheet['O'][int(cc3_num)+2].value = 'YES'
                    cc3_num+=1
                elif divide_sheet['B'][x].value[0] == '4':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #' + str(cc4_num)
                    basic_sheet['P'][int(cc4_num)+2].value = str_cc
                    basic_sheet['Q'][int(cc4_num)+2].value = divide_sheet['C'][x].value
                    basic_sheet['T'][int(cc4_num)+2].value = 'YES'
                    cc4_num+=1
                elif divide_sheet['B'][x].value[0] == '5':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #' + str(cc5_num)
                    basic_sheet['U'][int(cc5_num)+2].value = str_cc
                    basic_sheet['V'][int(cc5_num)+2].value = divide_sheet['C'][x].value
                    basic_sheet['Y'][int(cc5_num)+2].value = 'YES'
                    cc5_num+=1

            elif '[' in list(divide_sheet['C'][x].value):
                if divide_sheet['B'][x].value[0] == '1':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #M' + str(cc1_num2)
                    basic_sheet2['A'][int(cc1_num2)+2].value = str_cc
                    basic_sheet2['B'][int(cc1_num2)+2].value = divide_sheet['C'][x].value
                    basic_sheet2['E'][int(cc1_num2)+2].value = 'YES'
                    cc1_num2+=1
                elif divide_sheet['B'][x].value[0] == '2':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #M' + str(cc2_num2)
                    basic_sheet2['F'][int(cc2_num2)+2].value = str_cc
                    basic_sheet2['G'][int(cc2_num2)+2].value = divide_sheet['C'][x].value
                    basic_sheet2['J'][int(cc2_num2)+2].value = 'YES'
                    cc2_num2+=1
                elif divide_sheet['B'][x].value[0] == '3':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #M' + str(cc3_num2)
                    basic_sheet2['K'][int(cc3_num2)+2].value = str_cc
                    basic_sheet2['L'][int(cc3_num2)+2].value = divide_sheet['C'][x].value
                    basic_sheet2['O'][int(cc3_num2)+2].value = 'YES'
                    cc3_num2+=1
                elif divide_sheet['B'][x].value[0] == '4':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #M' + str(cc4_num2)
                    basic_sheet2['P'][int(cc4_num2)+2].value = str_cc
                    basic_sheet2['Q'][int(cc4_num2)+2].value = divide_sheet['C'][x].value
                    basic_sheet2['T'][int(cc4_num2)+2].value = 'YES'
                    cc4_num2+=1
                elif divide_sheet['B'][x].value[0] == '5':
                    str_cc = str(divide_sheet['B'][x].value[0]) + 'CC #M' + str(cc5_num2)
                    basic_sheet2['U'][int(cc5_num2)+2].value = str_cc
                    basic_sheet2['V'][int(cc5_num2)+2].value = divide_sheet['C'][x].value
                    basic_sheet2['Y'][int(cc5_num2)+2].value = 'YES'
                    cc5_num2+=1

        basic.save('SISO+MIMO_Divided.xlsx')

        self.Pviewer.append('SISO+MIMO completely divided !! \n')



    def okFunction(self):
        global send_name1
        global send_name2
        global send_name3
        global send_name4
        global data_siso
        global data_4b4
        global CC
        global rows

        data_siso=pd.read_excel(send_name1,sheet_name="Sheet1")
        data_4b4=pd.read_excel(send_name1,sheet_name="Sheet2")

        send=  op.load_workbook(send_name1)
        target1_Sheet = send['Sheet1']
        target1_Sheet_4b4 = send['Sheet2']

        target2 = op.load_workbook(send_name2)
        target2_Sheet = target2['Sheet1']
        target2_Sheet_4b4 = target2['Sheet2']
        #target2 = op.load_workbook(send_name3)
        #target1_Sheet = target1['Sheet1']

        CC=[]
        rows=[]

        def MakeSpace(v,k):
            global CC,rows

            for i in range(2,len(v.index)+2):
                vsp=v['CC.'+str(k)][i].split('-')

                temp=""
                for n in range(len(vsp)):
                    temp+=(' '+vsp[n])
                    if n<len(vsp)-1: temp+=('-')

                v['CC.'+str(k)][i]=temp


        def CheckSameCC (num,length,target_sheet):
            global CC,rows
            #length=len(d)
            Excelst=['A','F','K','P','U']

            for current_ca in range(length):
                for compare_ca in range(current_ca+1,length):
                    if CC[num]['CC.'+str(num)][current_ca+2]=='Same' or CC[num]['CC.'+str(num)][compare_ca+2]=='Same': continue
                    # 동일 CC 내에서도 Reverse 고려해야 하는지 질문
                    # if CC[num]['Reverse.'+str(num)][current_ca] == 'Yes' and CC[num]['Reverse.'+str(num)][compare_ca] == 'Yes':
                    current_split=CC[num]['CC.'+str(num)][current_ca+2].split('-')
                    compare_split=CC[num]['CC.'+str(num)][compare_ca+2].split('-')
                    compare_split_copy=copy.deepcopy(compare_split)

                    cnt=0

                    # 4X4 MIMO [ ]인 경우

                    for csp in current_split:
                        if csp in compare_split_copy:
                            cnt+=1

                            cspIndex=compare_split_copy.index(csp)
                            del compare_split_copy[cspIndex]

                    if cnt == len(current_split):

                        target_sheet[chr(ord(Excelst[num])+1)+str(compare_ca+4)].fill=PatternFill(fill_type='solid',start_color='808000')
                        target_sheet[chr(ord(Excelst[num])+3)+str(compare_ca+4)].fill=PatternFill(fill_type='solid',start_color='808000')

                        CC[num]['CC.'+str(num)][compare_ca+2]='Same'
                        CC[num]['Cover.'+str(num)][compare_ca+2]=CC[num]['Index.'+str(num)][current_ca+2]

                        target_sheet[chr(ord(Excelst[num])+1)+str(compare_ca+4)].value='Same'
                        target_sheet[chr(ord(Excelst[num])+3)+str(compare_ca+4)].value=CC[num]['Index.'+str(num)][current_ca+2]

        def ReadBasicTable(data,target_sheet):
            global data_siso,rows,CC

            for i in range(5):
                d=data.iloc[2:,i*5:(i+1)*5]
                d.columns.names=[str((i+2))+"CC Table"]
                d=d.dropna(how='all')
                rows.append(len(d))
                #d.drop_duplicates("CC."+str(i),keep="first",inplace=True)
                d.set_index=d['Index.'+str(i)]

                MakeSpace(d,i) #12A-2A 를 같은 2A로 인식하는데 이를 방지하기 위한 함수
                CC.append(d)
                CheckSameCC (i,len(d),target_sheet) #len(d)는 해당 cc의 row행의 수



        def CheckDiffCC(num): #num 0
            global CC,rows
            for compare_num in range(num+1,5):
                for left_index in range(2,len(CC[num])+2):    # left의 모든 행 수
                    if CC[num]['CC.'+str(num)][left_index]=='Same': continue

                    left_string=CC[num]['CC.'+str(num)][left_index]
                    left_reverse=CC[num]['Reverse.'+str(num)][left_index]
                    left=copy.deepcopy(CC[num]['CC.'+str(num)][left_index])

                    left=left.split('-')

                    signal=0
                    for right_index in range(2,len(CC[compare_num])+2): # right의 모든 행 수
                        if CC[compare_num]['CC.'+str(compare_num)][right_index]=='Same': continue

                        right_string=CC[compare_num]['CC.'+str(compare_num)][right_index]
                        right_reverse=CC[compare_num]['Reverse.'+str(compare_num)][right_index]

                        # 연속으로 데이터 비교가 가능할 때
                        if not (left_reverse=='No' and right_reverse=='No'):
                            if left_string in right_string:
                                CC[num]['Cover.'+str(num)][left_index]=CC[compare_num]['Index.'+str(compare_num)][right_index]
                                break

                            if left_string not in right_string and right_index<len(CC[compare_num])-1:
                                signal=1  # if in 수법으로 안통할때 signal ON
                                continue

                        if left_reverse=='No' and right_reverse=='No': signal=1

                        # 연속으로 데이터 비교가 불가능 할 때
                    if signal==1:
                        for right_index in range(2,len(CC[compare_num])+2):  #right 전체 행 수
                            #다시 재 setting 하고 RE AGAIN

                            left=copy.deepcopy(CC[num]['CC.'+str(num)][left_index])
                            left_reverse=CC[num]['Reverse.'+str(num)][left_index]
                            right=copy.deepcopy(CC[compare_num]['CC.'+str(compare_num)][right_index])
                            right_reverse=CC[compare_num]['Reverse.'+str(compare_num)][right_index]

                            left=left.split('-')
                            left_pcc=left[0]
                            right=right.split('-')
                            right_pcc=right[0]
                            cnt=1

                            if left_reverse=='No' and right_reverse=='No':
                                if left_pcc==right_pcc:
                                    left.pop(0)
                                    right.pop(0)

                            #if not (left_reverse=='No' and right_reverse=='No'):
                            while left:
                                lf=left.pop(0)
                                if lf in right:
                                    right.remove(lf)
                                    cnt+=1
                            if cnt!=len(CC[num]['CC.'+str(num)][left_index].split('-'))+1:
                                continue
                            if cnt==len(CC[num]['CC.'+str(num)][left_index].split('-'))+1:
                                CC[num]['Cover.'+str(num)][left_index]=CC[compare_num]['Index.'+str(compare_num)][right_index]
                                break

        def Toform (Rdata,reverse,target2_sheet,restriction=None):
            global index,CC,rows

            Rdata_split=Rdata.split('-')
            Rdata_list=copy.deepcopy(Rdata_split)
            for i in Rdata_split:
                if 'A' in i: continue
                if 'B' in i or 'C' in i: Rdata_list.append(i)
                if 'D' in i: Rdata_list.extend([i,i])
                if 'E' in i: Rdata_list.extend([i,i,i])


            if str(type(restriction))!="<class 'float'>":
                restriction_band=restriction.split()
                restriction_number=restriction_band[0][1:]

            # 5A-7A / 7A-5A
            FormBandScc=['E','F','G','H']
            #Rdata_set=list(set(Rdata_list))
            target2_sheet['A'+str(index)].value=Rdata


            # pcc 선언
            same_signal=0
            for pcc in range(len(Rdata_split)):
                if same_signal==1:
                    same_signal=0
                    continue
                Rdata_list_copy=copy.deepcopy(Rdata_list)
                Rdata_set=Rdata_split

                #그중 pcc는?
                #but restriction이 걸려있으면 예외처리를 해줘야지.
                if str(type(restriction))!="<class 'float'>":
                    if restriction_number==Rdata_set[pcc][:-1]:
                        continue

                target2_sheet['D'+str(index)].value=Rdata_set[pcc]

                where = Rdata_list_copy.index(Rdata_set[pcc])
                del Rdata_list_copy[where]
                for i,j in enumerate(Rdata_list_copy):
                    if 'A' in Rdata_set[pcc] and Rdata_set[pcc]==j:
                        same_signal=1

                    target2_sheet[FormBandScc[i]+str(index)].value=j

                index+=1

                if reverse=='No':
                    return
                if same_signal==1:
                    continue



#--------------------------엑셀 색깔 칠하는 코드 ------------------------------------------

        cover=['D','I','N','S','X']
        st=['A','F','K','P','U']
        stCC=['B','G','L','Q','V']
        global index
        index=6

        for idx,data in enumerate([data_siso,data_4b4]):
            if idx==0:
                ReadBasicTable(data,target1_Sheet)
                for num in range(5):
                    #0:1CC, 1:2CC, 2:3CC, 3:4CC 4:5CC
                    CheckDiffCC(num)

                for i in range(len(CC)):
                    for j in range(2,len(CC[i])+2):
                        #targetSheet[stCC[i]+str(store+2)].value=CC[i]['CC.'+str(i)][store]  # 중간에서 중복된 값들이 제거했을 때를 대비해서
                        target1_Sheet[cover[i]+str(j+2)].value=CC[i]['Cover.'+str(i)][j]

                        if str(type(target1_Sheet[cover[i]+str(j+2)].value)) != "<class 'float'>" and target1_Sheet[stCC[i]+str(j+2)].value != 'Same':
                            # cover 가능한 case
                            target1_Sheet[chr(ord(st[i])+1)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='666699')
                            target1_Sheet[chr(ord(st[i])+3)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='666699')

                        if str(type(target1_Sheet[cover[i]+str(j+2)].value)) == "<class 'float'>":
                            # 측정해야 할 case
                            target1_Sheet[chr(ord(st[i])+1)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='FFFF99')
                            target1_Sheet[chr(ord(st[i])+3)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='FFFF99')

                            Toform(target1_Sheet[chr(ord(st[i])+1)+str(j+2)].value,CC[i]['Reverse.'+str(i)][j],target2_Sheet,CC[i]['Restriction.'+str(i)][j])


                target1_Sheet.merge_cells('A1:Y1')
                target1_Sheet['A1']='LTE Release 10 Carrier Aggregation'
                target1_Sheet['A1'].font=Font(bold=True, size=16)
                target1_Sheet['A1'].alignment=Alignment(horizontal='center',vertical='center')

                self.Pviewer.append('[SISO]\n 측정해야 할 Downlink CA조합이 나왔습니다. \n')

                send.save("CA_table_Analyzed.xlsx")
                target2.save("DL_CA_combinations.xlsx")

            #4X4 MIMO
            if idx==1:
                CC=[]
                rows=[]

                ReadBasicTable(data,target1_Sheet_4b4)
                for num in range(5):
                    #0:1CC, 1:2CC, 2:3CC, 3:4CC 4:5CC
                    CheckDiffCC(num)

                index=6
                for i in range(len(CC)):
                    for j in range(2,len(CC[i])+2):
                        #targetSheet[stCC[i]+str(store+2)].value=CC[i]['CC.'+str(i)][store]  # 중간에서 중복된 값들이 제거했을 때를 대비해서
                        target1_Sheet_4b4[cover[i]+str(j+2)].value=CC[i]['Cover.'+str(i)][j]

                        if str(type(target1_Sheet_4b4[cover[i]+str(j+2)].value)) != "<class 'float'>" and target1_Sheet_4b4[stCC[i]+str(j+2)].value != 'Same':
                            # cover 가능한 case
                            target1_Sheet_4b4[chr(ord(st[i])+1)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='666699')
                            target1_Sheet_4b4[chr(ord(st[i])+3)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='666699')

                        if str(type(target1_Sheet_4b4[cover[i]+str(j+2)].value)) == "<class 'float'>":
                            # 측정해야 할 case
                            target1_Sheet_4b4[chr(ord(st[i])+1)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='FFFF99')
                            target1_Sheet_4b4[chr(ord(st[i])+3)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='FFFF99')

                            Toform(target1_Sheet_4b4[chr(ord(st[i])+1)+str(j+2)].value,CC[i]['Reverse.'+str(i)][j],target2_Sheet_4b4,CC[i]['Restriction.'+str(i)][j])


                target1_Sheet_4b4.merge_cells('A1:Y1')
                target1_Sheet_4b4['A1']='LTE Release 10 Carrier Aggregation'
                target1_Sheet_4b4['A1'].font=Font(bold=True, size=16)
                target1_Sheet_4b4['A1'].alignment=Alignment(horizontal='center',vertical='center')

                self.Pviewer.append('[4X4 MIMO]\n 측정해야 할 Downlink CA조합이 나왔습니다. \n')

                send.save("CA_table_Analyzed.xlsx")
                target2.save("DL_CA_combinations.xlsx")

        print("button2-----------FINISH------------------")


#=================================================================================================
    def okFunction2(self):
        global send_name1
        global send_name2
        global send_name3
        global send_name4
        global chBW
        global whole
        global lte
        global Frequency

        data2=pd.read_excel(send_name3,sheet_name="Sheet1")
        data2_4b4=pd.read_excel(send_name3,sheet_name="Sheet2")

        target3 = op.load_workbook(send_name4)

        target3_Sheet = target3['Sheet1']
        target3_Sheet_4b4 = target3['Sheet2']

        chBW=[1.4,3,5,10,15,20]
        lte="LTE"
        # 0: uplink Frequency
        # 1: downlink frequency
        # 2,3: 나도 정확히는 모르겠음. 무튼 channel 폭 계산할 때 사용 (uplink,downlink 순)
        # 4: 사용 가능한 bandwidth
        # 5: downlink low channel ( 여기에 [3] * 10 하면 == downlink high channel)
        # 6: uplink pcc 보고 downlink pcc 값 넣기 위해 추가 (sar팀 엑셀 파일 sheet2에 나와있는 밴드만 적용.)
        Frequency={'LTE1':[(1920,1980),(2110,2170),60,60,(range(2,6)),0],
               'LTE2':[(1850,1910),(1930,1990),60,60,(range(6)),600,18000],
               'LTE3':[(1710,1785),(1805,1880),75,75,(range(6)),1200],
               'LTE4':[(1710,1755),(2110,2155),45,45,(range(6)),1950,18000],
               'LTE5':[(824,849),(869,894),25,25,(range(4)),2400,18000],
               'LTE6':[(830,840),(875,885),10,10,(range(2,4)),2650],
               'LTE7':[(2500,2570),(2620,2690),70,70,(range(2,6)),2750,18000],
               'LTE8':[(880,915),(925,960),35,35,(range(4)),3450],
               'LTE9':[(1749.9,1784.9),(1844.9,1879.9),35,35,(range(2,6)),3800],
               'LTE10':[(1710,1770),(2110,2170),60,60,(range(2,6)),4150],
               'LTE11':[(1427.9,1447.9),(1475.9,1495.9),20,20,(range(2,4)),4750],
               'LTE12':[(699,716),(729,746),17,17,(range(4)),5010,18000],
               'LTE13':[(777,787),(746,756),10,10,(range(2,4)),5180,18000], #다름 (이해안감)
               'LTE14':[(788,798),(758,768),10,10,(range(2,4)),5280],
               'LTE17':[(704,716),(734,746),12,12,(range(2,4)),5730,18000], #?
               'LTE18':[(815,830),(860,875),15,15,(range(2,5)),5850],
               'LTE19':[(830,845),(875,890),15,15,(range(2,5)),6000],
               'LTE20':[(832,862),(791,821),30,30,(range(2,6)),6150],
               'LTE21':[(1447.9,1462.9),(1495.9,1510.9),15,15,(range(2,5)),6450],
               'LTE22':[(3410,3490),(3510,3590),80,80,(range(2,6)),6600],
               'LTE23':[(2000,2020),(2180,2200),20,20,(range(6)),7500],
               'LTE24':[(1626.5,1660.5),(1525,2559),34,34,(range(2,4)),7700],
               'LTE25':[(1850,1915),(1930,1995),65,65,(range(6)),8040,18000],
               'LTE26':[(814,849),(859,894),35,35,(range(5)),8690,18000],
               'LTE27':[(807,824),(859,869),17,17,(range(4)),9040,18000],#참고
               'LTE28':[(703,748),(758,803),45,45,(range(1,6)),9210],
               'LTE29':[(0,0),(717,728),0,11,range(4),9660],
               'LTE30':[(2305,2315),(2350,2360),10,10,(range(4)),9770,17890], #왜 bw 20이 아님?
               'LTE31':[(452.5,257.5),(462.5,467.5),5,5,(range(3)),9870],
               'LTE32':[(0,0),(1452,1496),0,44,9920],
               #LTE32: CHANNEL 불확실
               #---------------LTE 33-40 : TDD , else FDD-------------------
               'LTE33':[(1900,1920),(1900,1920),20,20,(range(2,6)),36000],
               'LTE34':[(2010,2025),(2010,2025),15,15,(range(2,5)),36200],
               'LTE35':[(1850,1910),(1850,1910),60,60,(range(6)),36350],
               'LTE36':[(1930,1990),(1930,1990),60,60,(range(6)),36950],
               'LTE37':[(1910,1930),(1910,1930),20,20,(range(2,6)),37550],
               'LTE38':[(2570,2620),(2570,2620),50,50,(range(2,6)),37750,0],
               'LTE39':[(1880,1920),(1880,1920),40,40,(range(2,6)),38250],
               'LTE40':[(2300,2400),(2300,2400),100,100,(range(2,4)),38650], #참고
               'LTE41':[(2496,2690),(2496,2690),194,194,(range(2,6)),39650,0],
               'LTE42':[(3400,3600),(3400,3600),200,200,(range(2,6)),41590],
               'LTE43':[(3600,3800),(3600,3800),200,200,(range(2,6)),43590],
               'LTE44':[(703,803),(703,803),100,100,(range(1,6)),45590],
               'LTE45':[(1447,1467),(1447,1467),20,20,(range(2,6)),46590],
               'LTE46':[(5150,5925),(5150,5925),775,775,(3,5),46790],
               'LTE47':[(5855,5925),(5855,5925),70,70,(3,5),54540],
               'LTE48':[(3550,3700),(3550,3700),150,150,(range(2,6)),55240],#?
               'LTE49':[(3550,3700),(3550,3700),150,150,(3,5),56740],
               'LTE50':[(1432,1517),(1432,1517),85,85,(range(1,6)),58240],
               'LTE51':[(1427,1432),(1427,1432),5,5,(range(1,3)),59090],
               'LTE52':[(3300,3400),(3300,3400),100,100,(range(2,6)),59140],
               'LTE65':[(1920,2010),(2110,2200),90,90,(range(6)),65536],
               #----------66 , 70 : FDD-------------------
               'LTE66':[(1710,1780),(2110,2180),70,70,(range(6)),66436,65536], #???
               'LTE67':[(0,0),(738,758),0,20,None,67336],
               'LTE68':[(698,728),(753,783),30,30,(range(2,5)),67536],
               'LTE69':[(0,0),(2570,2620),0,50,None,67836],
               'LTE70':[(1695,1710),(1995,2020),15,25,(range(2,6)),68336],
               'LTE71':[(663,698),(617,652),35,35,(range(2,6)),68586],
               'LTE72':[(451,456),(461,466),5,5,(range(3)),68936],
               'LTE73':[(450,455),(460,465),5,5,(range(3)),68986],
               'LTE74':[(1427,1470),(1475,1518),43,43,(range(6)),69036],
               'LTE75':[(0,0),(1432,1517),0,85,None,69466],
               'LTE76':[(0,0),(1427,1432),0,5,None,70316],
               'LTE85':[(698,716),(728,746),18,18,(range(2,4)),70366]
               }


        btn2_data=data2[4:][:].dropna(how='all')
        btn2_data_4b4 = data2_4b4[4:][:].dropna(how='all')

        global ULpcc,DLpcc
        ULpcc=['J','K','L','M']
        DLpcc=['N','O','P']

        def drop(value):
            value=value.strip('[')
            value=value.strip(']')
            return value


        def UL2DL(whole,target3_sheet,idx,ul,bw,chn,freq,rbo):
            global chBW,lte,Frequency,ULpcc,DLpcc

            ul=str(ul)
            ulstate=""
            if Frequency[lte+ul][0][0]+bw/2 == freq:
                ulstate="LOW"
            if (Frequency[lte+ul][0][0]+Frequency[lte+ul][0][1])/2 == freq:
                ulstate="MIDDLE"
            if Frequency[lte+ul][0][1]-bw/2 == freq:
                ulstate="HIGH"

            #
            #print("ulstate",ulstate)
            #bandwidth * 10 하면 channel 크기 조정

            target3_sheet[DLpcc[0]+str(idx+6)].value=bw
            if ulstate=="LOW":
                target3_sheet[DLpcc[1]+str(idx+6)].value= Frequency[lte+ul][5]+bw/2*10
                target3_sheet[DLpcc[2]+str(idx+6)].value=Frequency[lte+ul][1][0]+bw/2

            if ulstate=="MIDDLE":
                target3_sheet[DLpcc[1]+str(idx+6)].value=Frequency[lte+ul][5]+(Frequency[lte+ul][3]*10)/2
                target3_sheet[DLpcc[2]+str(idx+6)].value=(Frequency[lte+ul][1][0]+Frequency[lte+ul][1][1])/2

            if ulstate=="HIGH":
                target3_sheet[DLpcc[1]+str(idx+6)].value=Frequency[lte+ul][5]+(Frequency[lte+ul][3]*10)-bw/2*10
                target3_sheet[DLpcc[2]+str(idx+6)].value=Frequency[lte+ul][1][1]-bw/2

            return ulstate



        def CorrectPcc (whole,target_sheet,pcc,bw,chn,freq):
        #BW 올바르게 입력했는지 체크
            global chBW,lte,Frequency
            pcc=str(pcc)
            #if chBW[max(Frequency[lte+pcc][4])] != bw: return False
            # bandwidth가 20일수도 15일수도 다 자기멋대로일 때

            #else:
                #low, middle, high channel 중 어느 부분인지 확인하고, 그 값이 맞는지 확인
            #print("--------- ",bw,chn,freq)
            if Frequency[lte+pcc][5]+bw/2*10 == chn and Frequency[lte+pcc][1][0]+bw/2 == freq:
                return "LOW"
            elif Frequency[lte+pcc][5]+(Frequency[lte+pcc][3]*10)/2 == chn and (Frequency[lte+pcc][1][0]+Frequency[lte+pcc][1][1])/2 == freq:
                return "MIDDLE"
            elif Frequency[lte+pcc][5]+(Frequency[lte+pcc][3]*10)-bw/2*10 == chn and Frequency[lte+pcc][1][1]-bw/2 == freq:
                return "HIGH"
            return False

        global FormDLscc
        FormDLscc=['Q','T','W','Z']

        def WriteScc (whole,target3_sheet,scc,num,idx,state,Schn=None,Sfreq=None,exIdx=None,pccBW=None):  #num은 몇번째 scc인지
            global chBW,lte,Frequency
            global FormDLscc,Except
            scc=str(scc)
            #if scc=='66': print("state",state)

            if pccBW==None: bw=chBW[max(Frequency[lte+scc][4])]
            if pccBW!=None: bw=pccBW

            if state=="MIDDLE":
                target3_sheet[FormDLscc[num]+str(idx+6)].value = bw
                target3_sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value = Frequency[lte+scc][5]+(Frequency[lte+scc][3]*10)/2
                target3_sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value = (Frequency[lte+scc][1][0]+Frequency[lte+scc][1][1])/2
                return state

            elif state == "LOW":
                target3_sheet[FormDLscc[num]+str(idx+6)].value = bw
                target3_sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value = Frequency[lte+scc][5]+bw/2*10
                target3_sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value = Frequency[lte+scc][1][0]+(bw/2)
                return state

            elif state == "HIGH":
                target3_sheet[FormDLscc[num]+str(idx+6)].value = bw
                target3_sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value = Frequency[lte+scc][5]+(Frequency[lte+scc][3]*10)-(bw/2)*10
                target3_sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value = Frequency[lte+scc][1][1]-(bw/2)
                return state

            elif state == "C":
                target3_sheet[FormDLscc[num]+str(idx+6)].value = bw
                if num==3:
                    target3_sheet['AA'+str(idx+6)].value = Schn
                    target3_sheet['AB'+str(idx+6)].value = Sfreq
                    return state

                target3_sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value = Schn
                target3_sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value = Sfreq
                return state


        def Spacing(bw1,bw2):
            temp=(bw1+bw2-0.1*abs(bw1-bw2))/2
            v=0
            temp=temp/0.3
            temp=int(temp)
            v=temp*0.3
            return v

        def Intra(whole,target3_sheet,caseNum,state,pcc,pccBW,scc,num,idx,pidx=None):  #num 몇번재 scc, k 몇번째 인덱스
            global chBW,lte,Frequency
            global FormDLscc,Except
            pcc=str(pcc)
            scc=str(scc)
            pccState=state[0]
            #if pcc=='66': print("??",pccState)

            if caseNum==1: # PCC-SCC, non-contiguous ('A')
                if pccState=="LOW" or pccState=="MIDDLE":
                    state.append("HIGH")
                    WriteScc(whole,target3_sheet,scc,num,idx,"HIGH",None,None,None,pccBW)
                elif pccState=="HIGH":
                    state.append("LOW")
                    WriteScc(whole,target3_sheet,scc,num,idx,"LOW",None,None,None,pccBW)
                return

            elif caseNum==2: #SCC-SCC, non-contiguous ('A')
                state.append("HIGH")
                WriteScc(whole,target3_sheet,scc,num,idx,"HIGH",None,None,None,None)

            elif caseNum==3: #PCC-SCC, contiguous ('B','C', etc)
                gap=Spacing(chBW[max(Frequency[lte+scc][4])],chBW[max(Frequency[lte+scc][4])])
                #print(">>>>",gap)
                state.append("C")

                if pccState=="LOW":
                    pccfreq=Frequency[lte+pcc][1][0]+pccBW/2

                    # contiguous 하게 앞에다가? low freq 뒤에다가? high freq 붙일꺼니
                    lowfreq=pccfreq-gap
                    highfreq=pccfreq+gap
                    lowfreq_chn=Frequency[lte+pcc][5]+10*(lowfreq-Frequency[lte+pcc][1][0])
                    highfreq_chn=Frequency[lte+pcc][5]+10*(highfreq-Frequency[lte+pcc][1][0])

                    if lowfreq<Frequency[lte+pcc][1][0]:
                        freq=highfreq
                        chn=highfreq_chn
                    elif highfreq>Frequency[lte+pcc][1][1]:
                        freq=lowfreq
                        chn=lowfreq_chn
                    elif Frequency[lte+pcc][1][0]<lowfreq and Frequency[lte+pcc][1][1]>highfreq:
                        freq=highfreq
                        chn=highfreq_chn

                    WriteScc(whole,target3_sheet,scc,num,idx,"C",chn,freq,None,pccBW)

                elif pccState=="MIDDLE":
                    chn=Frequency[lte+pcc][5]+(Frequency[lte+pcc][3]*10)/2 + chBW[max(Frequency[lte+scc][4])]
                    freq=(Frequency[lte+pcc][1][0]+Frequency[lte+pcc][1][1])/2 + gap
                    WriteScc(whole,target3_sheet,scc,num,idx,"C",chn,freq,None,pccBW)

                elif pccState=="HIGH":
                    pccfreq=Frequency[lte+pcc][1][1]-pccBW/2

                    # contiguous 하게 앞에다가? low freq 뒤에다가? high freq 붙일꺼니
                    lowfreq=pccfreq-gap
                    highfreq=pccfreq+gap
                    lowfreq_chn=Frequency[lte+pcc][5]+10*(lowfreq-Frequency[lte+pcc][1][0])
                    highfreq_chn=Frequency[lte+pcc][5]+10*(highfreq-Frequency[lte+pcc][1][0])

                    if lowfreq<Frequency[lte+pcc][1][0]:
                        freq=highfreq
                        chn=highfreq_chn
                    elif highfreq>Frequency[lte+pcc][1][1]:
                        freq=lowfreq
                        chn=lowfreq_chn
                    elif Frequency[lte+pcc][1][0]<lowfreq and Frequency[lte+pcc][1][1]>highfreq:
                        freq=highfreq
                        chn=highfreq_chn

                    WriteScc(whole,target3_sheet,scc,num,idx,"C",chn,freq,None,pccBW)

            elif caseNum==4: #SCC-SCC, contiguous ('B','C', etc) ;; pidx 필요
                gap=Spacing(chBW[max(Frequency[lte+scc][4])],chBW[max(Frequency[lte+scc][4])])
                state.append("C")

                # 그냥 scc끼리만 contiguous 일 때
                # 처음 scc 상태를 보고 그걸 기준으로 앞에붙일지 뒤에 붙일지 결정
                # pcc까지 다같이 contiguous 일 때
                # 만약 두번째 scc일때는 이전 contiguous인 scc가 앞에붙였는지 뒤에 붙였는지 체크해야함.

                prevchn=target3_sheet[chr(ord(FormDLscc[pidx])+1)+str(idx+6)].value
                prevfreq=target3_sheet[chr(ord(FormDLscc[pidx])+2)+str(idx+6)].value

                lowfreq=prevfreq-gap
                highfreq=prevfreq+gap
                lowfreq_chn=prevchn - 10*gap
                highfreq_chn=prevchn + 10*gap

                if lowfreq<Frequency[lte+scc][1][0]:
                    freq=highfreq
                    chn=highfreq_chn
                elif highfreq>Frequency[lte+scc][1][1]:
                    freq=lowfreq
                    chn=lowfreq_chn
                elif Frequency[lte+scc][1][0]<lowfreq and highfreq<Frequency[lte+scc][1][1]:
                    freq=highfreq
                    chn=highfreq_chn


                WriteScc(whole,target3_sheet,scc,num,idx,"C",chn,freq,None,None)

        global Except
        Except=[['LTE66',15,67061,2172.5,"HIGH"]]
        #이건 이제 무시해도 되는 코드임

        for idx,whole in enumerate([btn2_data,btn2_data_4b4]):
            if idx==0:
                target3_sheet = target3_Sheet
                
            elif idx==1:
                target3_sheet = target3_Sheet_4b4

            for k in range(len(whole)):

                current=[]
                current_w=[]
                state=[]
                cnt=0

                pink=whole.iloc[k][3+cnt]  # E-UTRA CA configuration

                while pink:
                    pink=drop(pink)
                    current.append(pink)
                    cnt+=1
                    if str(type(whole.iloc[k][3+cnt])) == "<class 'float'>": break
                    if 3+cnt==8: break  #혹시 넘어갈까봐 bands pcc,scc 등
                    pink=whole.iloc[k][3+cnt]

                # frequency band # 뽑아내려고 분석하는 코드

                if current[-1] == " ": del current[-1]
                currentCopy=copy.deepcopy(current[1:]) #scc만 모아놓은 리스트
                currentCopy_set=set(currentCopy)
                for i in range(len(current)):
                    for j in range(len(current[i])):
                        if current[i][j] in ['A','B','C','D','E']:
                            current_w.append(current[i][j])
                            current[i]=current[i][:j]
                            break

                # 입력된 uplink를 토대로 downlink에 값을 넣는 방법
                ULpcc=['J','K','L','M']
                ULbw=target3_sheet[ULpcc[0]+str(k+6)].value
                ULchn=target3_sheet[ULpcc[1]+str(k+6)].value
                ULfreq=target3_sheet[ULpcc[2]+str(k+6)].value
                ULrbo=target3_sheet[ULpcc[3]+str(k+6)].value

                #uplink값만 넣었을 때, 실행해야하는 코드
                ulstate=UL2DL(whole,target3_sheet,k,current[0],ULbw,ULchn,ULfreq,ULrbo)

                # 입력한 downlink pcc가 올바른 값인지 확인하는 코드 n/o/p
                pcc=current[0]
                #RBO=CorrectPcc(int(pcc),whole.iloc[k][13],whole.iloc[k][14],whole.iloc[k][15])
                RBO=CorrectPcc(whole,target3_sheet,int(pcc),target3_sheet[DLpcc[0]+str(k+6)].value,target3_sheet[DLpcc[1]+str(k+6)].value,target3_sheet[DLpcc[2]+str(k+6)].value)

                exIdx=None
                if RBO == False:
                    #print("**** WRONG ****")
                    #print("\n")
                    #print(k,"번째")
                    #예외처리
                    if ulstate != RBO:  #(여기도 uplink 값만 넣었을 때 넣어야 하는 코드)
                        #print("**** UL PCC != DL PCC ***=")
                        print("")
                state.append(RBO)

            # SCC 값 넣어주는 코드
                sccNum=-1
                for i in range(1,len(current)): #PCC 제외한 나머지
                    # INTER BAND (PCC-SCC)
                    if pcc!=current[i]:
                        pccBW=target3_sheet[DLpcc[0]+str(k+6)].value
                        cntA=0
                        cntOther=0
                        where=sccNum
                        for idz,z in enumerate(current[:i]): # i 이전 꺼 검사
                            if current[i] == z:
                                if current_w[i]=='A' and current_w[idz]=='A':
                                # intra band , non contiguous
                                    cntA+=1
                                elif current_w[i] in ['B','C','D','E'] and  current_w[idz] in ['B','C','D','E']:
                                # intra band, contiguous
                                    cntOther+=1
                                    where=idz

                        if cntA==0 and cntOther==0: # INTER BAND 그 자체
                            sccNum+=1
                            #if RBO=="EXCEPT": WriteScc(current[i],sccNum,k,"MIDDLE",None,None,exIdx,None)

                            testCopy=[]
                            cnt=0
                            for x in current[1:]:
                                if x==current[i]:
                                    cnt+=1
                            if cnt>1 and current_w[i]=='A':
                                s=WriteScc(whole,target3_sheet,current[i],sccNum,k,"LOW",None,None,None,None)
                            else:s=WriteScc(whole,target3_sheet,current[i],sccNum,k,"MIDDLE",None,None,None,None)
                            state.append(s)

                        elif cntA>=1: # CASE 2: SCC-SCC non-contiguous
                            sccNum+=1
                            Intra(whole,target3_sheet,2,state,pcc,pccBW,current[i],sccNum,k,where-1)

                        elif cntOther>0: # CASE 4 : 진짜 only SCC간의 SCC-SCC contiguous
                            sccNum+=1
                            Intra(whole,target3_sheet,4,state,pcc,pccBW,current[i],sccNum,k,where-1)


                    # INTRA BAND (PCC-SCC) _ bandwidth가 중요
                    else: # CASE 3- PCC-SCC; contiguous
                        pccBW=target3_sheet[DLpcc[0]+str(k+6)].value  #20일수도 15일수도 있응게?

                        cntA=0
                        cntOther=0
                        for idz,z in enumerate(current[:i]): # i 이전 꺼 검사
                            if current[i] == z:
                                if current_w[i]=='A' and current_w[idz]=='A':
                                # intra band , non contiguous
                                    cntA+=1
                                elif current_w[i] in ['B','C','D','E'] and  current_w[idz] in ['B','C','D','E']:
                                # intra band, contiguous
                                    cntOther+=1  # pcc-scc가 같은 band에 존재할 뿐 아니라, scc간의 contiguous 조합도 존재한다.
                                    where=idz

                        if cntOther==0 and cntA==0:
                            sccNum+=1
                            Intra(whole,target3_sheet,1,state,pcc,pccBW,current[i],sccNum,k)

                        elif cntOther>0 and cntA==0: # CASE 3: PCC-SCC contiguous 뿐 아니라 이어서 SCC-SCC 간의 contiguous도 존재한다.
                            sccNum+=1
                            Intra(whole,target3_sheet,3,state,pcc,pccBW,current[i],sccNum,k) #PCC-SCC contiguous
                            #ntra(4,state,pcc,pccBW,current[i],sccNum,k,where-1) # PCC-SCC-SCC contiguous

                        elif cntA>0 and cntOther==0: # CASEv1 : PCC -SCC non contiguous
                            sccNum+=1
                            Intra(whole,target3_sheet,1,state,pcc,pccBW,current[i],sccNum,k)

        self.Pviewer.append('Downlink CA Output Power 시트가 완성되었습니다. \n')
        target3.save('Final_Downlink_CA_output_power.xlsx')
        print("okbutton3-----------FINISH------------------")


if __name__ == '__main__':
   app = QApplication(sys.argv)
   ex = MyApp()
   sys.exit(app.exec_())
