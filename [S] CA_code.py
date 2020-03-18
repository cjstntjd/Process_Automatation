import openpyxl as op
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
import copy
#Index #NCC #Restriction #Covered #Reverse
data = pd.read_excel("./basicTable.xlsx")
data2 = pd.read_excel("./formTable.xlsx")

target1_Ad="targetTable.xlsx"
target2_Ad="formTable.xlsx"
target3_Ad="resultTable.xlsx"

target1 = op.load_workbook(target1_Ad)
target1_Sheet = target1['Sheet1']
target2 = op.load_workbook(target2_Ad)
target2_Sheet = target2['Sheet1']
target3 = op.load_workbook(target3_Ad)
target3_Sheet = target3['Sheet1']


def MakeSpace(v,k):
    for i in range(2,len(v.index)+2):
        vsp=v['CC.'+str(k)][i].split('-')

        temp=""
        for n in range(len(vsp)):
            temp+=(' '+vsp[n])
            if n<len(vsp)-1: temp+=('-')

        v['CC.'+str(k)][i]=temp


def CheckSameCC (num,length):
    #length=len(d)
    Excelst=['A','F','K','P']
    for current_ca in range(length):
        for compare_ca in range(current_ca+1,length):
            if CC[num]['CC.'+str(num)][current_ca+2]=='Same' or CC[num]['CC.'+str(num)][compare_ca+2]=='Same': continue
            # 동일 CC 내에서도 Reverse 고려해야 하는지 질문
            # if CC[num]['Reverse.'+str(num)][current_ca] == 'Yes' and CC[num]['Reverse.'+str(num)][compare_ca] == 'Yes':
            current_split=CC[num]['CC.'+str(num)][current_ca+2].split('-')
            compare_split=CC[num]['CC.'+str(num)][compare_ca+2].split('-')
            compare_split_copy=copy.deepcopy(compare_split)

            cnt=0
            for csp in current_split:
                if csp in compare_split_copy:
                    cnt+=1
                    cspIndex=compare_split_copy.index(csp)
                    del compare_split_copy[cspIndex]

            if cnt == len(current_split):

                target1_Sheet[chr(ord(Excelst[num])+1)+str(compare_ca+4)].fill=PatternFill(fill_type='solid',start_color='99CC00')
                target1_Sheet[chr(ord(Excelst[num])+3)+str(compare_ca+4)].fill=PatternFill(fill_type='solid',start_color='99CC00')

                CC[num]['CC.'+str(num)][compare_ca+2]='Same'
                CC[num]['Cover.'+str(num)][compare_ca+2]=CC[num]['Index.'+str(num)][current_ca+2]

                target1_Sheet[chr(ord(Excelst[num])+1)+str(compare_ca+4)].value='Same'
                target1_Sheet[chr(ord(Excelst[num])+3)+str(compare_ca+4)].value=CC[num]['Index.'+str(num)][current_ca+2]


CC=[]
rows=[]
def ReadBasicTable():
    global data,rows
    #Max_CC = 4 #직접 UI에서 입력하도록 (조합하고자 하는 파일에 나열된 CC의 종류가 최대 몇 CC인지)
    for i in range(5-1):
        d=data.iloc[2:,i*5:(i+1)*5]
        d.columns.names=[str((i+2))+"CC Table"]
        d=d.dropna(how='all')
        rows.append(len(d))
        #d.drop_duplicates("CC."+str(i),keep="first",inplace=True)
        d.set_index=d['Index.'+str(i)]

        MakeSpace(d,i)
        CC.append(d)
        CheckSameCC (i,len(d))


def CheckDiffCC(num): #num 0
    for compare_num in range(num+1,4):
        for left_index in range(2,len(CC[num])+2):    # left의 모든 행 수
            if CC[num]['CC.'+str(num)][left_index]=='Same': continue

            left_string=CC[num]['CC.'+str(num)][left_index]
            left_reverse=CC[num]['Reverse.'+str(num)][left_index]
            left=copy.deepcopy(CC[num]['CC.'+str(num)][left_index])

            left=left.split('-')

            signal=0
            for right_index in range(2,len(CC[compare_num])+2): # right의 모든 행 수
                if CC[compare_num]['CC.'+str(compare_num)][right_index]=='Same': continue

                right_string=CC[compare_num]['CC.'+str(compare_num)][right_index]
                right_reverse=CC[compare_num]['Reverse.'+str(compare_num)][right_index]

                # 연속으로 데이터 비교가 가능할 때
                if not (left_reverse=='No' and right_reverse=='No'):
                    if left_string in right_string:
                        CC[num]['Cover.'+str(num)][left_index]=CC[compare_num]['Index.'+str(compare_num)][right_index]
                        break

                    if left_string not in right_string and right_index<len(CC[compare_num])-1:
                        signal=1  # if in 수법으로 안통할때 signal ON
                        continue

                if left_reverse=='No' and right_reverse=='No': signal=1

                # 연속으로 데이터 비교가 불가능 할 때
            if signal==1:
                for right_index in range(2,len(CC[compare_num])+2):  #right 전체 행 수
                    #다시 재 setting 하고 RE AGAIN

                    left=copy.deepcopy(CC[num]['CC.'+str(num)][left_index])
                    left_reverse=CC[num]['Reverse.'+str(num)][left_index]
                    right=copy.deepcopy(CC[compare_num]['CC.'+str(compare_num)][right_index])
                    right_reverse=CC[compare_num]['Reverse.'+str(compare_num)][right_index]

                    left=left.split('-')
                    left_pcc=left[0]
                    right=right.split('-')
                    right_pcc=right[0]
                    cnt=1

                    if left_reverse=='No' and right_reverse=='No':
                        if left_pcc==right_pcc:
                            left.pop(0)
                            right.pop(0)

                    #if not (left_reverse=='No' and right_reverse=='No'):
                    while left:
                        lf=left.pop(0)
                        if lf in right:
                            right.remove(lf)
                            cnt+=1
                    if cnt!=len(CC[num]['CC.'+str(num)][left_index].split('-'))+1:
                        continue
                    if cnt==len(CC[num]['CC.'+str(num)][left_index].split('-'))+1:
                        CC[num]['Cover.'+str(num)][left_index]=CC[compare_num]['Index.'+str(compare_num)][right_index]
                        break

ReadBasicTable()
for num in range(4):
    #0:2CC, 1:3CC, 2:4CC, 3:5CC
    CheckDiffCC(num)

def Toform (Rdata):
    global index

    Rdata_split=Rdata.split('-')
    Rdata_list=copy.deepcopy(Rdata_split)
    for i in Rdata_split:
        if 'A' in i: continue
        if 'B' in i or 'C' in i: Rdata_list.append(i)
        if 'D' in i: Rdata_list.extend([i,i])
        if 'E' in i: Rdata_list.extend([i,i,i])

    # 5A-7A / 7A-5A
    FormBandScc=['E','F','G','H']
    Rdata_set=list(set(Rdata_list))
    Rdata_set.sort()  #그냥 보기 좋게 하려고 sort() 함.

    target2_Sheet['A'+str(index)].value=Rdata

    for pcc in range(len(Rdata_set)):
        Rdata_list_copy=copy.deepcopy(Rdata_list)
        Rdata_set=list(set(Rdata_list))
        Rdata_set.sort()  #그냥 보기 좋게 하려고 sort() 함.
        target2_Sheet['D'+str(index)].value=Rdata_set[pcc]

        where = Rdata_list_copy.index(Rdata_set[pcc])
        del Rdata_list_copy[where]
        for i,j in enumerate(Rdata_list_copy):
            target2_Sheet[FormBandScc[i]+str(index)].value=j

        index+=1



cover=['D','I','N','S']
st=['A','F','K','P']
stCC=['B','G','L','Q']

index=6
for i in range(len(CC)):
    for j in range(2,len(CC[i])+2):
        #targetSheet[stCC[i]+str(store+2)].value=CC[i]['CC.'+str(i)][store]  # 중간에서 중복된 값들이 제거했을 때를 대비해서
        target1_Sheet[cover[i]+str(j+2)].value=CC[i]['Cover.'+str(i)][j]

        if str(type(target1_Sheet[cover[i]+str(j+2)].value)) != "<class 'float'>" and target1_Sheet[stCC[i]+str(j+2)].value != 'Same':
            # cover 가능한 case
            target1_Sheet[chr(ord(st[i])+1)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='A9A9A9')
            target1_Sheet[chr(ord(st[i])+3)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='A9A9A9')

        if str(type(target1_Sheet[cover[i]+str(j+2)].value)) == "<class 'float'>":
            # 측정해야 할 case
            target1_Sheet[chr(ord(st[i])+1)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='ADFF2F')
            target1_Sheet[chr(ord(st[i])+3)+str(j+2)].fill=PatternFill(fill_type='solid',start_color='ADFF2F')

            Toform(target1_Sheet[chr(ord(st[i])+1)+str(j+2)].value)


target1_Sheet.merge_cells('A1:T1')
target1_Sheet['A1']='LTE Release 10 Carrier Aggregation'
target1_Sheet['A1'].font=Font(bold=True, size=16)
target1_Sheet['A1'].alignment=Alignment(horizontal='center',vertical='center')

target1.save(target1_Ad)
target2.save(target2_Ad)

#-------------------------------------------- form에 실제 데이터 옮기기 --------------------------
chBW=[1.4,3,5,10,15,20]

Frequency={'LTE1':[(1920,1980),(2110,2170),60,60,(range(2,6)),0],
           'LTE2':[(1850,1910),(1930,1990),60,60,(range(6)),600],
           'LTE3':[(1710,1785),(1805,1880),75,75,(range(6)),1200],
           'LTE4':[(1710,1755),(2110,2155),45,45,(range(6)),1950],
           'LTE5':[(824,849),(869,894),25,25,(range(4)),2400],
           'LTE6':[(830,840),(875,885),10,10,(range(2,4)),2650],
           'LTE7':[(2500,2570),(2620,2690),70,70,(range(2,6)),2750],
           'LTE8':[(880,915),(925,960),35,35,(range(4)),3450],
           'LTE9':[(1749.9,1784.9),(1844.9,1879.9),35,35,(range(2,6)),3800],
           'LTE10':[(1710,1770),(2110,2170),60,60,(range(2,6)),4150],
           'LTE11':[(1427.9,1447.9),(1475.9,1495.9),20,20,(range(2,4)),4750],
           'LTE12':[(699,716),(729,746),17,17,(range(4)),5010],
           'LTE13':[(777,787),(746,756),10,10,(range(2,4)),5180], #다름 (이해안감)
           'LTE14':[(788,798),(758,768),10,10,(range(2,4)),5280],
           'LTE17':[(704,716),(734,746),12,12,(range(2,4)),5730], #?
           'LTE18':[(815,830),(860,875),15,15,(range(2,5)),5850],
           'LTE19':[(830,845),(875,890),15,15,(range(2,5)),6000],
           'LTE20':[(832,862),(791,821),30,30,(range(2,6)),6150],
           'LTE21':[(1447.9,1462.9),(1495.9,1510.9),15,15,(range(2,5)),6450],
           'LTE22':[(3410,3490),(3510,3590),80,80,(range(2,6)),6600],
           'LTE23':[(2000,2020),(2180,2200),20,20,(range(6)),7500],
           'LTE24':[(1626.5,1660.5),(1525,2559),34,34,(range(2,4)),7700],
           'LTE25':[(1850,1915),(1930,1995),65,65,(range(6)),8040],
           'LTE26':[(814,849),(859,894),35,35,(range(5)),8690],
           'LTE27':[(807,824),(859,869),17,17,(range(4)),9040],#참고
           'LTE28':[(703,748),(758,803),45,45,(range(1,6)),9210],
           'LTE29':[(0,0),(717,728),0,11,range(4),9660],
           'LTE30':[(2305,2315),(2350,2360),10,10,(range(4)),9770], #왜 bw 20이 아님?
           'LTE31':[(452.5,257.5),(462.5,467.5),5,5,(range(3)),9870],
           'LTE32':[(0,0),(1452,1496),0,44,9920],
           #LTE32: CHANNEL 불확실
           #LTE 33-40 : TDD , else FDD
           'LTE33':[(1900,1920),(1900,1920),20,20,(range(2,6)),36000],
           'LTE34':[(2010,2025),(2010,2025),15,15,(range(2,5)),36200],
           'LTE35':[(1850,1910),(1850,1910),60,60,(range(6)),36350],
           'LTE36':[(1930,1990),(1930,1990),60,60,(range(6)),36950],
           'LTE37':[(1910,1930),(1910,1930),20,20,(range(2,6)),37550],
           'LTE38':[(2570,2620),(2570,2620),50,50,(range(2,6)),37750],
           'LTE39':[(1880,1920),(1880,1920),40,40,(range(2,6)),38250],
           'LTE40':[(2300,2400),(2300,2400),100,100,(range(2,4)),38650], #참고
           'LTE41':[(2496,2690),(2496,2690),194,194,(range(2,6)),39650],
           'LTE42':[(3400,3600),(3400,3600),200,200,(range(2,6)),41590],
           'LTE43':[(3600,3800),(3600,3800),200,200,(range(2,6)),43590],
           'LTE44':[(703,803),(703,803),100,100,(range(1,6)),45590],
           'LTE45':[(1447,1467),(1447,1467),20,20,(range(2,6)),46590],
           'LTE46':[(5150,5925),(5150,5925),775,775,(3,5),46790],
           'LTE47':[(5855,5925),(5855,5925),70,70,(3,5),54540],
           'LTE48':[(3550,3700),(3550,3700),150,150,(range(2,6)),55240],#?
           'LTE49':[(3550,3700),(3550,3700),150,150,(3,5),56740],
           'LTE50':[(1432,1517),(1432,1517),85,85,(range(1,6)),58240],
           'LTE51':[(1427,1432),(1427,1432),5,5,(range(1,3)),59090],
           'LTE52':[(3300,3400),(3300,3400),100,100,(range(2,6)),59140],
           'LTE65':[(1920,2010),(2110,2200),90,90,(range(6)),65536],
           'LTE66':[(1710,1780),(2110,2180),70,70,(range(6)),66436],
           'LTE67':[(0,0),(738,758),0,20,None,67336],
           'LTE68':[(698,728),(753,783),30,30,(range(2,5)),67536],
           'LTE69':[(0,0),(2570,2620),0,50,None,67836],
           'LTE70':[(1695,1710),(1995,2020),15,25,(range(2,6)),68336],
           'LTE71':[(663,698),(617,652),35,35,(range(2,6)),68586],
           'LTE72':[(451,456),(461,466),5,5,(range(3)),68936],
           'LTE73':[(450,455),(460,465),5,5,(range(3)),68986],
           'LTE74':[(1427,1470),(1475,1518),43,43,(range(6)),69036],
           'LTE75':[(0,0),(1432,1517),0,85,None,69466],
           'LTE76':[(0,0),(1427,1432),0,5,None,70316],
           'LTE85':[(698,716),(728,746),18,18,(range(2,4)),70366]
           }

whole=data2[4:][:].dropna(how='all')
lte="LTE"

def CorrectPcc (pcc,bw,chn,freq):
    #BW 올바르게 입력했는지 체크
    global chBW,lte
    pcc=str(pcc)
    if chBW[max(Frequency[lte+pcc][4])] != bw: return False
    else:
        #low, middle, high channel 중 어느 부분인지 확인하고, 그 값이 맞는지 확인
        if Frequency[lte+pcc][5]+bw/2*10 == chn and Frequency[lte+pcc][1][0]+bw/2 == freq:
            return "LOW"
        elif Frequency[lte+pcc][5]+(Frequency[lte+pcc][3]*10)/2 == chn and (Frequency[lte+pcc][1][0]+Frequency[lte+pcc][1][1])/2 == freq:
            return "MIDDLE"
        elif Frequency[lte+pcc][5]+(Frequency[lte+pcc][3]*10)-bw/2*10 == chn and Frequency[lte+pcc][1][1]-bw/2 == freq:
            return "HIGH"
        return False

FormDLscc=['Q','T','W','Z']

def WriteScc (scc,num,idx,state,Schn=None,Sfreq=None,exIdx=None):  #num은 몇번째 scc인지
    global FormDLscc,Except
    scc=str(scc)

    bw=chBW[max(Frequency[lte+scc][4])]

    if state == "EXCEPT":
        target3_Sheet[FormDLscc[num]+str(idx+6)].value = Except[exIdx][1]
        target3_Sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value =  Schn
        target3_Sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value =  Sfreq
        return state

    elif state=="MIDDLE":
        target3_Sheet[FormDLscc[num]+str(idx+6)].value = bw
        target3_Sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value = Frequency[lte+scc][5]+(Frequency[lte+scc][3]*10)/2
        target3_Sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value = (Frequency[lte+scc][1][0]+Frequency[lte+scc][1][1])/2
        return state

    elif state == "LOW":
        target3_Sheet[FormDLscc[num]+str(idx+6)].value = bw
        target3_Sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value = Frequency[lte+scc][5]+bw/2*10
        target3_Sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value = Frequency[lte+scc][1][0]+(bw/2)
        return state

    elif state == "HIGH":
        target3_Sheet[FormDLscc[num]+str(idx+6)].value = bw
        target3_Sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value = Frequency[lte+scc][5]+(Frequency[lte+scc][3]*10)-(bw/2)*10
        target3_Sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value = Frequency[lte+scc][1][1]-(bw/2)
        return state

    elif state == "C":
        target3_Sheet[FormDLscc[num]+str(idx+6)].value = bw
        target3_Sheet[chr(ord(FormDLscc[num])+1)+str(idx+6)].value = Schn
        target3_Sheet[chr(ord(FormDLscc[num])+2)+str(idx+6)].value = Sfreq
        return state


def Spacing(bw1,bw2):
    temp=bw1+bw2-0.1*abs(bw1-bw2)
    v=0
    while v<=temp:
        v+=0.3
    v-=0.3
    return v

def Intra(caseNum,state,pcc,scc,num,idx,pidx=None,exIdx=None):  #num 몇번재 scc, k 몇번째 인덱스
    global FormDLscc,Except
    pcc=str(pcc)
    scc=str(scc)
    pccState=state[0]

    if caseNum==1: # PCC-SCC, non-contiguous ('A')
        if pccState=="LOW" or pccState=="MIDDLE":
            state.append("HIGH")
            WriteScc(scc,num,idx,"HIGH")
        elif pccState=="HIGH":
            state.append("LOW")
            WriteScc(scc,num,idx,"LOW")
        elif pccState=="EXCEPT":
            if Except[exIdx][4] == "LOW" or Except[exIdx][4]=="MIDDLE":
                state.append("HIGH")
                WriteScc(scc,num,idx,"HIGH",None,None,exIdx)
            elif Except[exIdx][4] == "HIGH":
                state.append("LOW")
                WriteScc(scc,num,idx,"LOW",None,None,exIdx)

        return

    elif caseNum==2: #SCC-SCC, non-contiguous ('A')
        state.append("HIGH")
        WriteScc(scc,num,idx,"HIGH")

    elif caseNum==3: #PCC-SCC, contiguous ('B','C', etc)
        gap=Spacing(chBW[max(Frequency[lte+scc][4])],chBW[max(Frequency[lte+scc][4])])
        state.append("C")

        if pccState == "EXCEPT":
            gap = Spacing(Except[exIdx][1],Except[exIdx][1])
            chn = Except[exIdx][2] + Except[exIdx][1]
            freq = Except[exIdx][3] + gap
            if freq > Frequency[lte+pcc][1][1]-(chBW[max(Frequency[lte+pcc][4])]/2):
                chn = Except[exIdx][2] - Except[exIdx][1]
                freq = Except[exIdx][3] - gap
            WriteScc(scc,num,idx,"EXCEPT",chn,freq,exIdx)
            return

        elif pccState=="LOW":
            chn=Frequency[lte+pcc][5]+chBW[max(Frequency[lte+pcc][4])]/2*10 + chBW[max(Frequency[lte+scc][4])]
            freq=Frequency[lte+pcc][1][0]+Frequency[lte+pcc][3]/2 + gap
            WriteScc(scc,num,idx,"C",chn,freq)

        elif pccState=="MIDDLE":
            chn=Frequency[lte+pcc][5]+(Frequency[lte+pcc][3]*10)/2 + chBW[max(Frequency[lte+scc][4])]
            freq=(Frequency[lte+pcc][1][0]+Frequency[lte+pcc][1][1])/2 + gap
            WriteScc(scc,num,idx,"C",chn,freq)

        elif pccState=="HIGH":
            chn=Frequency[lte+pcc][5]+(Frequency[lte+pcc][3]*10)-(chBW[max(Frequency[lte+pcc][4])]/2)*10 + chBW[max(Frequency[lte+scc][4])]
            freq=Frequency[lte+pcc][1][1]-(chBW[max(Frequency[lte+pcc][4])]/2) + gap
            if freq > Frequency[lte+pcc][1][1]-(chBW[max(Frequency[lte+pcc][4])]/2):
                freq=Frequency[lte+pcc][1][1]-(chBW[max(Frequency[lte+pcc][4])]/2) - gap
                chn=Frequency[lte+pcc][5]+(Frequency[lte+pcc][3]*10)-(chBW[max(Frequency[lte+pcc][4])]/2)*10 - chBW[max(Frequency[lte+scc][4])]
            WriteScc(scc,num,idx,"C",chn,freq)

    elif caseNum==4: #SCC-SCC, contiguous ('B','C', etc) ;; pidx 필요
        gap=Spacing(chBW[max(Frequency[lte+scc][4])],chBW[max(Frequency[lte+scc][4])])
        state.append("C")
        chn=target3_Sheet[chr(ord(FormDLscc[pidx])+1)+str(idx+6)].value + chBW[max(Frequency[lte+scc][4])]
        freq=target3_Sheet[chr(ord(FormDLscc[pidx])+2)+str(idx+6)].value + gap
        if freq > Frequency[lte+scc][1][1]-(chBW[max(Frequency[lte+scc][4])]/2):
            freq=target3_Sheet[chr(ord(FormDLscc[pidx])+2)+str(idx+6)].value - gap
            chn=target3_Sheet[chr(ord(FormDLscc[pidx])+1)+str(idx+6)].value - chBW[max(Frequency[lte+scc][4])]
        WriteScc(scc,num,idx,"C",chn,freq)


Except=[['LTE66',15,67061,2172.5,"HIGH"]]

for k in range(len(whole)):
    current=[]
    current_w=[]
    state=[]
    cnt=0

    pink=whole.iloc[k][3+cnt]
    while pink:
        current.append(pink)
        cnt+=1
        if str(type(whole.iloc[k][3+cnt])) == "<class 'float'>": break
        pink=whole.iloc[k][3+cnt]

    # frequency band # 뽑아내려고 분석하는 코드
    if current[-1] == " ": del current[-1]
    for i in range(len(current)):
        for j in range(len(current[i])):
            if current[i][j] in ['A','B','C','D','E']:
                current_w.append(current[i][j])
                current[i]=current[i][:j]
                break

    # 입력한 downlink pcc가 올바른 값인지 확인하는 코드 n/o/p
    pcc=current[0]
    RBO=CorrectPcc(int(pcc),whole.iloc[k][13],whole.iloc[k][14],whole.iloc[k][15])
    exIdx=None
    if RBO == False:
        print("**** WRONG ****")
        #예외처리
        for aId,a in enumerate(Except):
            if a[0]==lte+pcc: #LTE 2343
                if whole.iloc[k][14] == a[2] and whole.iloc[k][15] == a[3]:
                     RBO="EXCEPT"
                     exIdx=aId
                     print("aid",aId,a)
                     break
    state.append(RBO)

    # SCC 값 넣어주는 코드
    sccNum=-1
    for i in range(1,len(current)): #PCC 제외한 나머지
        # INTER BAND (PCC-SCC)
        if pcc!=current[i]:
            cntA=0
            cntOther=0
            where=sccNum
            for idz,z in enumerate(current[:i]): # i 이전 꺼 검사
                if current[i] == z:
                    if current_w[i]=='A' and current_w[idz]=='A':
                    # intra band , non contiguous
                        cntA+=1
                    elif current_w[i] in ['B','C','D','E'] and  current_w[idz] in ['B','C','D','E']:
                    # intra band, contiguous
                        cntOther+=1
                        where=idz

            if cntA==0 and cntOther==0: # INTER BAND 그 자체
                sccNum+=1
                if RBO=="EXCEPT": WriteScc(current[i],sccNum,k,"MIDDLE",None,None,exIdx)
                s=WriteScc(current[i],sccNum,k,"MIDDLE")
                state.append(s)

            elif cntA>=1: # CASE 2: SCC-SCC non-contiguous
                sccNum+=1
                if RBO=="EXCEPT":
                    Intra(2,state,pcc,current[i],sccNum,k,None,exIdx)
                    continue
                Intra(2,state,pcc,current[i],sccNum,k)

            elif cntOther>0: # CASE 4 : SCC-SCC contiguous
                sccNum+=1
                if RBO=="EXCEPT":
                     Intra(4,state,pcc,current[i],sccNum,k,where-1,exIdx)
                     continue
                Intra(4,state,pcc,current[i],sccNum,k,where-1)



        # INTRA BAND (PCC-SCC)
        else: # CASE 3- PCC-SCC; contiguous
            cntA=0
            cntOther=0
            for idz,z in enumerate(current[:i]): # i 이전 꺼 검사
                if current[i] == z:
                    if current_w[i]=='A' and current_w[idz]=='A':
                    # intra band , non contiguous
                        cntA+=1
                    elif current_w[i] in ['B','C','D','E'] and  current_w[idz] in ['B','C','D','E']:
                    # intra band, contiguous
                        cntOther+=1
                        where=idz

            if cntOther==0 and cntA==0:
                sccNum+=1
                if RBO=="EXCEPT":
                    Intra(1,state,pcc,current[i],sccNum,k,None,exIdx)
                    continue
                Intra(1,state,pcc,current[i],sccNum,k)

            elif cntOther>0: # CASE 3: PCC-SCC contiguous
                sccNum+=1
                if RBO=="EXCEPT":
                    Intra(3,state,pcc,current[i],sccNum,k,None,exIdx)
                    continue
                Intra(3,state,pcc,current[i],sccNum,k)

            elif cntA>0: # CASEv1 : PCC -SCC non contiguous
                sccNum+=1
                if RBO=="EXCEPT":
                    Intra(1,state,pcc,current[i],sccNum,k,None,exIdx)
                    continue
                Intra(1,state,pcc,current[i],sccNum,k)



target3.save(target3_Ad)
