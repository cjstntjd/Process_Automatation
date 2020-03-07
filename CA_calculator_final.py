import openpyxl as op
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
import copy
#Index #NCC #Restriction #Covered #Reverse
data = pd.read_excel("./exdata4.xlsx")
targetAd="copy6.xlsx"
target = op.load_workbook(targetAd)
targetSheet = target['Sheet1']
ResultSheet = target.create_sheet("Result")
l=["Index","CC","Reverse"]
cnt=1
ResultSheet.merge_cells('B1:D1')
for id,i in enumerate(['A','B','E']):
    ResultSheet[i+str(1)]=l[id]
    ResultSheet[i+str(1)].font=Font(bold=True)
    ResultSheet[i+str(1)].alignment=Alignment(horizontal='center',vertical='center')
    if id==0: ResultSheet[i+str(1)].fill=PatternFill(patternType='solid',start_color='87CEEB')
    if id==1: ResultSheet[i+str(1)].fill=PatternFill(patternType='solid',start_color='FFB6C1')
    if id==2: ResultSheet[i+str(1)].fill=PatternFill(patternType='solid',start_color='9370DB')

def Space(v,k):
    for i in range(2,len(v.index)+2):
        vsp=v['CC.'+str(k)][i].split('-')

        temp=""
        for n in range(len(vsp)):
            temp+=(' '+vsp[n])
            if n<len(vsp)-1: temp+=('-')

        v['CC.'+str(k)][i]=temp
CC=[]
rows=[]
def Search():
    global data,rows
    #Max_CC = 4 #직접 UI에서 입력하도록 (조합하고자 하는 파일에 나열된 CC의 종류가 최대 몇 CC인지)
    for i in range(5-1):
        d=data.iloc[2:,i*5:(i+1)*5]
        d.columns.names=[str((i+2))+"CC Table"]
        d=d.dropna(how='all')
        rows.append(len(d))
        d.drop_duplicates("CC."+str(i),keep="first",inplace=True)
        d.set_index=d['Index.'+str(i)]

        Space(d,i)
        CC.append(d)


def Step1(num): #num 0
    for k in range(num+1,5-1):   # (2,3) (2,4) (3,4)
        for now in range(2,len(CC[num])+2):    # left의 모든 행 수
            left_string=CC[num]['CC.'+str(num)][now]
            left_reverse=CC[num]['Reverse.'+str(num)][now]
            left=copy.deepcopy(CC[num]['CC.'+str(num)][now])

            left=left.split('-')
            #left_pcc=left[0]
            signal=0
            for w in range(2,len(CC[k])+2): # right의 모든 행 수
                right_string=CC[k]['CC.'+str(k)][w]
                right_reverse=CC[k]['Reverse.'+str(k)][w]
                #right=right.split('-')
                #right_pcc=right[0]

                # 연속으로 데이터 비교가 가능할 때
                if not (left_reverse=='No' and right_reverse=='No'):
                    if left_string in right_string:
                        CC[num]['Cover.'+str(num)][now]=CC[k]['Index.'+str(k)][w]
                        break

                    if left_string not in right_string and w<len(CC[k])-1:
                        signal=1  # if in 수법으로 안통할때 signal ON
                        continue

                if left_reverse=='No' and right_reverse=='No': signal=1

                # 연속으로 데이터 비교가 불가능 할 때
            if signal==1:
                for w in range(2,len(CC[k])+2):  #right 전체 행 수
                    #다시 재 setting 하고 RE AGAIN
                    left=copy.deepcopy(CC[num]['CC.'+str(num)][now])
                    left_reverse=CC[num]['Reverse.'+str(num)][now]
                    right=copy.deepcopy(CC[k]['CC.'+str(k)][w])
                    right_reverse=CC[k]['Reverse.'+str(k)][w]

                    left=left.split('-')
                    left_pcc=left[0]
                    right=right.split('-')
                    right_pcc=right[0]
                    cnt=1

                    if left_reverse=='No' and right_reverse=='No':
                        if left_pcc==right_pcc:
                            left.pop(0)
                            right.pop(0)

                    #if not (left_reverse=='No' and right_reverse=='No'):
                    while left:
                        lf=left.pop(0)
                        if lf in right:
                            right.remove(lf)
                            cnt+=1
                    if cnt!=len(CC[num]['CC.'+str(num)][now].split('-'))+1:
                        continue
                    if cnt==len(CC[num]['CC.'+str(num)][now].split('-'))+1:
                        CC[num]['Cover.'+str(num)][now]=CC[k]['Index.'+str(k)][w]
                        break

Search()
for num in range(2,5):
    Step1(num-2)

cover=['D','I','N','S']
st=['A','F','K','P']
stCC=['B','G','L','Q']

for i in range(len(CC)):
    for store in range(2,len(CC[i])+2):
        targetSheet[stCC[i]+str(store+2)].value=CC[i]['CC.'+str(i)][store]  # 중간에서 중복된 값들이 제거했을 때를 대비해서
        targetSheet[cover[i]+str(store+2)].value=CC[i]['Cover.'+str(i)][store]
        if str(type(targetSheet[cover[i]+str(store+2)].value)) != "<class 'float'>":
            for a in range(4):  # 눈에 띄게 색칠하려고
                targetSheet[chr(ord(st[i])+a)+str(store+2)].fill=PatternFill(fill_type='solid',start_color='A9A9A9')

        if str(type(targetSheet[cover[i]+str(store+2)].value)) == "<class 'float'>":
            for a in range(4):  # 눈에 띄게 색칠하려고
                targetSheet[chr(ord(st[i])+a)+str(store+2)].fill=PatternFill(fill_type='solid',start_color='ADFF2F')
                if a<2:
                    print(chr(ord('A')+a)+str(cnt+1))
                    ResultSheet[chr(ord('A')+a)+str(cnt+1)].value=CC[i][l[a]+'.'+str(i)][store]
                if a==2:
                    ResultSheet[chr(ord('A')+a+2)+str(cnt+1)].value=CC[i][l[a]+'.'+str(i)][store]
            cnt+=1

    if len(CC[i])<rows[i]: #중복되고 남은 부분(output excel에서)
        print(len(CC[i]),rows[i])
        for g in range(rows[i]-len(CC[i])):
            for a in range(5):
                print(">>",targetSheet[chr(ord(st[i])+a)+str(len(CC[i])+2+g+2)].value)
                targetSheet[chr(ord(st[i])+a)+str(len(CC[i])+2+g+2)].value=" "

targetSheet.merge_cells('A1:T1')
targetSheet['A1']='LTE Release 10 Carrier Aggregation'
targetSheet['A1'].font=Font(bold=True, size=16)
targetSheet['A1'].alignment=Alignment(horizontal='center',vertical='center')

target.save(targetAd)
